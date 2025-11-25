import requests
from bs4 import BeautifulSoup, Tag
import json
import re
import time
from typing import Dict, Any, List, Optional
from urllib.parse import urljoin

def clean_text(text: str) -> str:
    """Removes excessive whitespace and cleans up text."""
    if not text:
        return ""
    return re.sub(r'\s+', ' ', text).strip()

def extract_basic_info(soup: BeautifulSoup) -> Dict[str, Any]:
    """Extract basic faculty information from profile page."""
    info = {
        'name': '',
        'title': '',
        'department': '',
        'email': '',
        'phone': '',
        'office': '',
        'address': '',
        'social_links': {}
    }
    
    # Name - usually in h1 or main heading
    name_elem = soup.find('h1')
    if not name_elem:
        name_elem = soup.find('h2')
    if name_elem:
        info['name'] = clean_text(name_elem.get_text())
    
    # Title and Department - usually near the name
    title_elem = soup.find(text=re.compile(r'Professor|Associate Professor|Assistant Professor|Lecturer'))
    if title_elem:
        info['title'] = clean_text(title_elem)
        
        # Department is often in the same area
        parent = title_elem.find_parent()
        if parent:
            dept_text = parent.get_text()
            dept_match = re.search(r'Department of ([^\n]+)', dept_text)
            if dept_match:
                info['department'] = clean_text(dept_match.group(1))
    
    # Contact information
    # Email
    email_elem = soup.find('a', href=re.compile(r'mailto:'))
    if email_elem:
        info['email'] = email_elem.get('href', '').replace('mailto:', '').strip()
    
    # Phone - look for "Ext." or phone pattern
    phone_text = soup.find(text=re.compile(r'Ext\.|Phone:|Tel:|\+880|\d{3}[-\s]\d{3}[-\s]\d{4}', re.IGNORECASE))
    if phone_text:
        phone_parent = phone_text.find_parent()
        if phone_parent:
            info['phone'] = clean_text(phone_parent.get_text())
    
    # Office - look for "Office:" or "Room"
    office_elem = soup.find(text=re.compile(r'Office:|Room:', re.IGNORECASE))
    if office_elem:
        office_parent = office_elem.find_parent()
        if office_parent:
            info['office'] = clean_text(office_parent.get_text())
    
    # Address - look for "Address:" section
    address_heading = soup.find(text=re.compile(r'Address:', re.IGNORECASE))
    if address_heading:
        address_parent = address_heading.find_parent()
        if address_parent:
            # Get the next sibling or content
            address_div = address_parent.find_next_sibling()
            if address_div:
                info['address'] = clean_text(address_div.get_text())
            else:
                info['address'] = clean_text(address_parent.get_text())
    
    # Social links (Google Scholar, LinkedIn, etc.)
    social_links = soup.find_all('a', href=re.compile(r'scholar.google|linkedin|researchgate|orcid', re.IGNORECASE))
    for link in social_links:
        href = link.get('href', '')
        if 'scholar.google' in href:
            info['social_links']['google_scholar'] = href
        elif 'linkedin' in href:
            info['social_links']['linkedin'] = href
        elif 'researchgate' in href:
            info['social_links']['researchgate'] = href
        elif 'orcid' in href:
            info['social_links']['orcid'] = href
    
    return info

def extract_accordion_sections(soup: BeautifulSoup) -> Dict[str, Any]:
    """Extract all expandable/accordion sections from the profile."""
    sections = {}
    
    # Common section titles that appear in faculty profiles
    section_patterns = [
        'Academic Background',
        'Short Biography',
        'Professional Experience',
        'Research Interest',
        'Selected Publications',
        'Publications',
        'Awards & Achievements',
        'Awards and Achievements',
        'Teaching Materials',
        'Courses Taught',
        'Affiliation Professional Membership',
        'Professional Membership',
        'Photo Gallery',
        'Certifications',
        'Skills',
        'Projects'
    ]
    
    # Look for accordion/collapsible elements
    # These are often in divs with classes like 'accordion', 'panel', 'collapse'
    accordion_items = soup.find_all(['div', 'section'], class_=re.compile(r'accordion|panel|collapse|tab'))
    
    # Also look for headings that might indicate sections
    headings = soup.find_all(['h2', 'h3', 'h4', 'h5'], text=re.compile('|'.join(section_patterns), re.IGNORECASE))
    
    # Process accordion items
    for item in accordion_items:
        # Find the heading/title of this section
        title_elem = item.find(['h2', 'h3', 'h4', 'h5', 'button', 'a'])
        if title_elem:
            title = clean_text(title_elem.get_text())
            
            # Get the content - usually in a div after or within the item
            content_div = item.find(['div', 'p'], class_=re.compile(r'content|body|text|collapse'))
            if not content_div:
                content_div = item
            
            content = clean_text(content_div.get_text())
            
            if title and content and len(content) > 10:
                # Normalize the title
                title_key = title.lower().replace(' ', '_').replace('&', 'and')
                sections[title_key] = content
    
    # Process headings that weren't in accordion items
    for heading in headings:
        title = clean_text(heading.get_text())
        title_key = title.lower().replace(' ', '_').replace('&', 'and')
        
        # Skip if we already have this section
        if title_key in sections:
            continue
        
        # Get all content until the next heading
        content_parts = []
        current = heading.find_next_sibling()
        
        while current:
            # Stop at next heading
            if current.name in ['h2', 'h3', 'h4', 'h5']:
                break
            
            if isinstance(current, Tag):
                text = clean_text(current.get_text())
                if text:
                    content_parts.append(text)
            
            current = current.find_next_sibling()
        
        if content_parts:
            sections[title_key] = ' '.join(content_parts)
    
    return sections

def scrape_faculty_profile(url: str, headers: Dict[str, str]) -> Optional[Dict[str, Any]]:
    """Scrape a single faculty member's profile page."""
    print(f"  → Scraping profile: {url}")
    
    try:
        resp = requests.get(url, headers=headers, timeout=30)
        resp.raise_for_status()
        soup = BeautifulSoup(resp.content, 'html.parser')
    except requests.exceptions.RequestException as e:
        print(f"    ✗ Error fetching profile: {e}")
        return None
    
    # Extract all information
    profile_data = {
        'profile_url': url,
        'basic_info': extract_basic_info(soup),
        'sections': extract_accordion_sections(soup)
    }
    
    # Get name for logging
    name = profile_data['basic_info'].get('name', 'Unknown')
    print(f"    ✓ Extracted profile for: {name}")
    print(f"      Sections found: {len(profile_data['sections'])}")
    
    return profile_data

def get_faculty_links(main_url: str, headers: Dict[str, str]) -> List[str]:
    """Get all faculty profile links from the main faculty page."""
    print(f"Fetching faculty members list from: {main_url}")
    
    try:
        resp = requests.get(main_url, headers=headers, timeout=30)
        resp.raise_for_status()
        soup = BeautifulSoup(resp.content, 'html.parser')
    except requests.exceptions.RequestException as e:
        print(f"✗ Error fetching main page: {e}")
        return []
    
    profile_links = []
    
    # Look for links that contain faculty member profiles
    # These usually have patterns like /faculty-members/, /profile/, or similar
    all_links = soup.find_all('a', href=True)
    
    for link in all_links:
        href = link.get('href', '')
        
        # Check if this looks like a faculty profile link
        if any(pattern in href.lower() for pattern in ['faculty', 'profile', 'member']):
            # Convert relative URLs to absolute
            full_url = urljoin(main_url, href)
            
            # Avoid duplicates and non-profile pages
            if full_url not in profile_links and full_url != main_url:
                # Additional check: profile links usually have a specific structure
                # and don't end with just the base URL
                if len(full_url.split('/')) > 5:
                    profile_links.append(full_url)
    
    # Alternative: Look for specific elements that contain faculty cards/items
    faculty_cards = soup.find_all(['div', 'article'], class_=re.compile(r'faculty|member|staff|person|profile', re.IGNORECASE))
    
    for card in faculty_cards:
        link = card.find('a', href=True)
        if link:
            href = link.get('href', '')
            full_url = urljoin(main_url, href)
            if full_url not in profile_links and full_url != main_url:
                profile_links.append(full_url)
    
    print(f"✓ Found {len(profile_links)} faculty profile links")
    return profile_links

def scrape_all_faculty(main_url: str, delay: float = 1.0) -> List[Dict[str, Any]]:
    """Scrape all faculty members from the department."""
    headers = {
        'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36'
    }
    
    # Get all profile links
    profile_links = get_faculty_links(main_url, headers)
    
    if not profile_links:
        print("✗ No faculty profiles found")
        return []
    
    # Scrape each profile
    all_profiles = []
    
    print(f"\n{'='*60}")
    print(f"Scraping {len(profile_links)} faculty profiles...")
    print(f"{'='*60}\n")
    
    for i, link in enumerate(profile_links, 1):
        print(f"[{i}/{len(profile_links)}]")
        
        profile_data = scrape_faculty_profile(link, headers)
        
        if profile_data:
            all_profiles.append(profile_data)
        
        # Be polite - add delay between requests
        if i < len(profile_links):
            time.sleep(delay)
    
    return all_profiles

def save_as_csv(faculty_data: List[Dict[str, Any]], filename: str):
    """Save faculty data as CSV format - flattened structure."""
    import csv
    
    if not faculty_data:
        return
    
    print(f"  → Saving CSV format to: {filename}")
    
    # Collect all possible section keys
    all_section_keys = set()
    for profile in faculty_data:
        all_section_keys.update(profile.get('sections', {}).keys())
    
    # Define CSV headers
    headers = [
        'name', 'title', 'department', 'email', 'phone', 'office', 
        'address', 'profile_url', 'google_scholar', 'linkedin', 
        'researchgate', 'orcid'
    ]
    
    # Add section columns
    section_headers = sorted(all_section_keys)
    headers.extend(section_headers)
    
    with open(filename, 'w', encoding='utf-8', newline='') as f:
        writer = csv.DictWriter(f, fieldnames=headers)
        writer.writeheader()
        
        for profile in faculty_data:
            row = {}
            
            # Basic info
            basic = profile.get('basic_info', {})
            row['name'] = basic.get('name', '')
            row['title'] = basic.get('title', '')
            row['department'] = basic.get('department', '')
            row['email'] = basic.get('email', '')
            row['phone'] = basic.get('phone', '')
            row['office'] = basic.get('office', '')
            row['address'] = basic.get('address', '')
            row['profile_url'] = profile.get('profile_url', '')
            
            # Social links
            social = basic.get('social_links', {})
            row['google_scholar'] = social.get('google_scholar', '')
            row['linkedin'] = social.get('linkedin', '')
            row['researchgate'] = social.get('researchgate', '')
            row['orcid'] = social.get('orcid', '')
            
            # Sections
            sections = profile.get('sections', {})
            for section_key in section_headers:
                row[section_key] = sections.get(section_key, '')
            
            writer.writerow(row)
    
    print(f"    ✓ CSV saved successfully")

def save_as_markdown(faculty_data: List[Dict[str, Any]], filename: str):
    """Save faculty data as Markdown format - human-readable documentation."""
    
    if not faculty_data:
        return
    
    print(f"  → Saving Markdown format to: {filename}")
    
    with open(filename, 'w', encoding='utf-8') as f:
        # Write header
        f.write("# EWU CSE Department Faculty Members\n\n")
        f.write(f"*Total Faculty: {len(faculty_data)}*\n\n")
        f.write("---\n\n")
        
        # Write each faculty profile
        for i, profile in enumerate(faculty_data, 1):
            basic = profile.get('basic_info', {})
            sections = profile.get('sections', {})
            
            name = basic.get('name', 'Unknown')
            title = basic.get('title', '')
            
            # Faculty header
            f.write(f"## {i}. {name}\n\n")
            
            if title:
                f.write(f"**{title}**\n\n")
            
            # Basic information table
            f.write("### Contact Information\n\n")
            f.write("| Field | Details |\n")
            f.write("|-------|--------|\n")
            
            if basic.get('department'):
                f.write(f"| Department | {basic['department']} |\n")
            if basic.get('email'):
                f.write(f"| Email | {basic['email']} |\n")
            if basic.get('phone'):
                f.write(f"| Phone | {basic['phone']} |\n")
            if basic.get('office'):
                f.write(f"| Office | {basic['office']} |\n")
            if basic.get('address'):
                f.write(f"| Address | {basic['address']} |\n")
            if profile.get('profile_url'):
                f.write(f"| Profile | [View Profile]({profile['profile_url']}) |\n")
            
            f.write("\n")
            
            # Social links
            social = basic.get('social_links', {})
            if social:
                f.write("### Online Presence\n\n")
                for platform, url in social.items():
                    platform_name = platform.replace('_', ' ').title()
                    f.write(f"- **{platform_name}**: [{url}]({url})\n")
                f.write("\n")
            
            # Sections
            if sections:
                for section_key, content in sections.items():
                    section_title = section_key.replace('_', ' ').title()
                    f.write(f"### {section_title}\n\n")
                    f.write(f"{content}\n\n")
            
            f.write("---\n\n")
    
    print(f"    ✓ Markdown saved successfully")

def save_as_excel(faculty_data: List[Dict[str, Any]], filename: str):
    """Save faculty data as Excel format with multiple sheets."""
    try:
        import openpyxl
        from openpyxl.styles import Font, Alignment, PatternFill
        from openpyxl.utils import get_column_letter
    except ImportError:
        print("    ⚠ openpyxl not installed. Skipping Excel export.")
        print("    Install with: pip install openpyxl")
        return
    
    if not faculty_data:
        return
    
    print(f"  → Saving Excel format to: {filename}")
    
    # Create workbook
    wb = openpyxl.Workbook()
    
    # Remove default sheet
    if 'Sheet' in wb.sheetnames:
        wb.remove(wb['Sheet'])
    
    # Sheet 1: Summary
    ws_summary = wb.create_sheet("Faculty Summary")
    
    # Headers
    headers = ['#', 'Name', 'Title', 'Department', 'Email', 'Phone', 'Office', 'Profile URL']
    ws_summary.append(headers)
    
    # Style header row
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    
    for col in range(1, len(headers) + 1):
        cell = ws_summary.cell(1, col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Data rows
    for i, profile in enumerate(faculty_data, 1):
        basic = profile.get('basic_info', {})
        row = [
            i,
            basic.get('name', ''),
            basic.get('title', ''),
            basic.get('department', ''),
            basic.get('email', ''),
            basic.get('phone', ''),
            basic.get('office', ''),
            profile.get('profile_url', '')
        ]
        ws_summary.append(row)
    
    # Auto-adjust column widths
    for col in range(1, len(headers) + 1):
        max_length = 0
        column = get_column_letter(col)
        for cell in ws_summary[column]:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws_summary.column_dimensions[column].width = adjusted_width
    
    # Sheet 2: Detailed Information
    ws_detailed = wb.create_sheet("Detailed Profiles")
    
    # Write detailed information
    current_row = 1
    
    for profile in faculty_data:
        basic = profile.get('basic_info', {})
        sections = profile.get('sections', {})
        
        # Name header
        name_cell = ws_detailed.cell(current_row, 1)
        name_cell.value = basic.get('name', 'Unknown')
        name_cell.font = Font(bold=True, size=14)
        name_cell.fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
        current_row += 1
        
        # Basic info
        info_items = [
            ('Title', basic.get('title', '')),
            ('Department', basic.get('department', '')),
            ('Email', basic.get('email', '')),
            ('Phone', basic.get('phone', '')),
            ('Office', basic.get('office', '')),
            ('Address', basic.get('address', '')),
        ]
        
        for label, value in info_items:
            if value:
                ws_detailed.cell(current_row, 1).value = label
                ws_detailed.cell(current_row, 1).font = Font(bold=True)
                ws_detailed.cell(current_row, 2).value = value
                current_row += 1
        
        current_row += 1
        
        # Sections
        for section_key, content in sections.items():
            section_title = section_key.replace('_', ' ').title()
            
            section_cell = ws_detailed.cell(current_row, 1)
            section_cell.value = section_title
            section_cell.font = Font(bold=True, color="366092")
            current_row += 1
            
            content_cell = ws_detailed.cell(current_row, 1)
            content_cell.value = content
            content_cell.alignment = Alignment(wrap_text=True, vertical='top')
            ws_detailed.merge_cells(start_row=current_row, start_column=1, 
                                   end_row=current_row, end_column=5)
            current_row += 2
        
        current_row += 2  # Space between profiles
    
    # Adjust column widths for detailed sheet
    ws_detailed.column_dimensions['A'].width = 25
    ws_detailed.column_dimensions['B'].width = 60
    
    # Save workbook
    wb.save(filename)
    print(f"    ✓ Excel saved successfully")

def main():
    """Main function to orchestrate the faculty scraping."""
    
    print("="*60)
    print("EWU CSE FACULTY MEMBERS SCRAPER")
    print("="*60)
    print()
    
    # Main faculty page URL
    faculty_url = "https://fse.ewubd.edu/computer-science-engineering/faculty-members"
    
    # Scrape all faculty
    faculty_data = scrape_all_faculty(faculty_url, delay=1.0)
    
    # Print summary
    print("\n" + "="*60)
    print("✅ SCRAPING COMPLETE!")
    print("="*60)
    print(f"\n📊 Total faculty members scraped: {len(faculty_data)}")
    
    if faculty_data:
        print(f"\n📋 Faculty members:")
        for profile in faculty_data:
            name = profile['basic_info'].get('name', 'Unknown')
            title = profile['basic_info'].get('title', '')
            sections = len(profile['sections'])
            print(f"  • {name}")
            print(f"    Title: {title}")
            print(f"    Sections: {sections}")
            if profile['sections']:
                section_names = ', '.join(profile['sections'].keys())
                if len(section_names) > 80:
                    section_names = section_names[:77] + '...'
                print(f"    Data: {section_names}")
            print()
    
    # Save in multiple formats
    print("\n" + "="*60)
    print("💾 SAVING DATA IN MULTIPLE FORMATS")
    print("="*60)
    print()
    
    # Format 1: JSON (original)
    json_filename = 'ewu_cse_faculty_complete.json'
    print(f"1. JSON Format")
    with open(json_filename, 'w', encoding='utf-8') as f:
        json.dump(faculty_data, f, ensure_ascii=False, indent=2)
    print(f"   ✓ Saved to: {json_filename}")
    print()
    
    # Format 2: CSV
    csv_filename = 'ewu_cse_faculty_complete.csv'
    print(f"2. CSV Format")
    save_as_csv(faculty_data, csv_filename)
    print()
    
    # Format 3: Markdown
    md_filename = 'ewu_cse_faculty_complete.md'
    print(f"3. Markdown Format")
    save_as_markdown(faculty_data, md_filename)
    print()
    
    # Format 4: Excel (bonus)
    excel_filename = 'ewu_cse_faculty_complete.xlsx'
    print(f"4. Excel Format (bonus)")
    save_as_excel(faculty_data, excel_filename)
    print()
    
    print("="*60)
    print("✅ ALL FORMATS SAVED SUCCESSFULLY!")
    print("="*60)
    print("\n📄 Output Files:")
    print(f"  1. {json_filename} - Full structured data (JSON)")
    print(f"  2. {csv_filename} - Spreadsheet format (CSV)")
    print(f"  3. {md_filename} - Human-readable docs (Markdown)")
    print(f"  4. {excel_filename} - Excel workbook with multiple sheets")
    print("="*60)

if __name__ == '__main__':
    main()